#!/usr/bin/env python3
"""Validate stored MSFT financial rows against archived Microsoft FinancialStatement XLSX files."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

TICKER = "MSFT"
DEFAULT_COMPANY_PATH = Path("data/hyperscaler/companies/MSFT.json")
DEFAULT_XLSX_DIR = Path("data/hyperscaler/ir_documents/originals/MSFT")
DEFAULT_OUTPUT_DIR = Path("data/hyperscaler/validation/msft_financialstatement_xlsx")

MONEY_TOLERANCE = 1_000_000.0
SHARE_TOLERANCE = 1_000_000.0
PER_SHARE_TOLERANCE = 0.005


@dataclass(frozen=True)
class XlsxMetric:
    value: float | None
    method: str


@dataclass(frozen=True)
class MetricDef:
    metric: str
    statement: str
    unit_kind: str
    comparable: bool = True
    note: str = ""


BALANCE_METRICS = [
    MetricDef("total_assets", "balance_sheet", "money"),
    MetricDef("current_assets", "balance_sheet", "money"),
    MetricDef("cash_and_equivalents", "balance_sheet", "money"),
    MetricDef("inventory", "balance_sheet", "money"),
    MetricDef("current_investments", "balance_sheet", "money"),
    MetricDef("trade_and_non_trade_receivables", "balance_sheet", "money"),
    MetricDef("non_current_assets", "balance_sheet", "money"),
    MetricDef(
        "property_plant_and_equipment",
        "balance_sheet",
        "money",
        comparable=False,
        note="Financial Datasets and MSFT XLSX differ on whether operating lease ROU assets are included in some annual periods.",
    ),
    MetricDef("goodwill_and_intangible_assets", "balance_sheet", "money"),
    MetricDef(
        "investments",
        "balance_sheet",
        "money",
        comparable=False,
        note="Provider definition is broader than short-term investments plus the MSFT equity/other investments line in several periods.",
    ),
    MetricDef(
        "non_current_investments",
        "balance_sheet",
        "money",
        comparable=False,
        note="Provider definition is broader than the MSFT equity/other investments line in several periods.",
    ),
    MetricDef(
        "outstanding_shares",
        "balance_sheet",
        "shares",
        comparable=False,
        note="MSFT XLSX balance sheet gives rounded shares outstanding in millions; source may retain a more precise XBRL value.",
    ),
    MetricDef("total_liabilities", "balance_sheet", "money"),
    MetricDef("current_liabilities", "balance_sheet", "money"),
    MetricDef("current_debt", "balance_sheet", "money"),
    MetricDef("trade_and_non_trade_payables", "balance_sheet", "money"),
    MetricDef("deferred_revenue", "balance_sheet", "money"),
    MetricDef("deposit_liabilities", "balance_sheet", "money", note="Mapped to short-term plus long-term unearned revenue for MSFT."),
    MetricDef("non_current_liabilities", "balance_sheet", "money"),
    MetricDef("non_current_debt", "balance_sheet", "money"),
    MetricDef(
        "tax_liabilities",
        "balance_sheet",
        "money",
        comparable=False,
        note="Provider tax liability mapping is not consistently the same as MSFT long-term income taxes.",
    ),
    MetricDef("shareholders_equity", "balance_sheet", "money"),
    MetricDef("retained_earnings", "balance_sheet", "money"),
    MetricDef("accumulated_other_comprehensive_income", "balance_sheet", "money"),
    MetricDef("total_debt", "balance_sheet", "money"),
]

INCOME_METRICS = [
    MetricDef("revenue", "income_statement", "money"),
    MetricDef("cost_of_revenue", "income_statement", "money"),
    MetricDef("gross_profit", "income_statement", "money"),
    MetricDef("operating_expense", "income_statement", "money"),
    MetricDef("selling_general_and_administrative_expenses", "income_statement", "money"),
    MetricDef("research_and_development", "income_statement", "money"),
    MetricDef("operating_income", "income_statement", "money"),
    MetricDef("income_tax_expense", "income_statement", "money"),
    MetricDef("net_income", "income_statement", "money"),
    MetricDef("net_income_common_stock", "income_statement", "money"),
    MetricDef("preferred_dividends_impact", "income_statement", "money"),
    MetricDef("consolidated_income", "income_statement", "money"),
    MetricDef("earnings_per_share", "income_statement", "per_share"),
    MetricDef("earnings_per_share_diluted", "income_statement", "per_share"),
    MetricDef("weighted_average_shares", "income_statement", "shares"),
    MetricDef("weighted_average_shares_diluted", "income_statement", "shares"),
]

CASH_FLOW_METRICS = [
    MetricDef("net_income", "cash_flow_statement", "money"),
    MetricDef(
        "depreciation_and_amortization",
        "cash_flow_statement",
        "money",
        comparable=False,
        note="Microsoft XLSX line is 'Depreciation, amortization, and other', so it is not patched into the narrower source field.",
    ),
    MetricDef("share_based_compensation", "cash_flow_statement", "money"),
    MetricDef("net_cash_flow_from_operations", "cash_flow_statement", "money"),
    MetricDef("capital_expenditure", "cash_flow_statement", "money", note="XLSX value uses cash outflow sign from 'Additions to property and equipment'."),
    MetricDef("business_acquisitions_and_disposals", "cash_flow_statement", "money"),
    MetricDef("investment_acquisitions_and_disposals", "cash_flow_statement", "money", note="Purchases + maturities + sales of investments."),
    MetricDef("net_cash_flow_from_investing", "cash_flow_statement", "money"),
    MetricDef("issuance_or_repayment_of_debt_securities", "cash_flow_statement", "money"),
    MetricDef("issuance_or_purchase_of_equity_shares", "cash_flow_statement", "money"),
    MetricDef("dividends_and_other_cash_distributions", "cash_flow_statement", "money", note="XLSX value uses cash outflow sign from dividends paid."),
    MetricDef("net_cash_flow_from_financing", "cash_flow_statement", "money"),
    MetricDef("change_in_cash_and_equivalents", "cash_flow_statement", "money"),
    MetricDef("effect_of_exchange_rate_changes", "cash_flow_statement", "money"),
    MetricDef("ending_cash_balance", "cash_flow_statement", "money"),
    MetricDef("free_cash_flow", "cash_flow_statement", "money", note="Calculated as net cash from operations minus absolute additions to property and equipment."),
]

METRICS = BALANCE_METRICS + INCOME_METRICS + CASH_FLOW_METRICS
PATCHABLE_METRICS = [metric for metric in METRICS if metric.comparable]


def normalize_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ").replace("\n", " ")).strip().lower()


def worksheet_rows(sheet: Worksheet) -> list[tuple[str, str, tuple[Any, ...]]]:
    return [
        (normalize_label(row[0]), str(row[0]).replace("\n", " / "), row)
        for row in sheet.iter_rows(values_only=True)
        if row and row[0] is not None
    ]


def find_exact(sheet: Worksheet, label: str, col: int = 2) -> tuple[Any | None, str]:
    target = normalize_label(label)
    for normalized, original, row in worksheet_rows(sheet):
        if normalized == target:
            return row[col - 1], original
    return None, ""


def find_startswith(sheet: Worksheet, prefix: str, col: int = 2) -> tuple[Any | None, str]:
    target = normalize_label(prefix)
    for normalized, original, row in worksheet_rows(sheet):
        if normalized.startswith(target):
            return row[col - 1], original
    return None, ""


def find_first_exact(sheet: Worksheet, labels: list[str], col: int = 2) -> tuple[Any | None, str]:
    for label in labels:
        value, original = find_exact(sheet, label, col)
        if value is not None:
            return value, original
    return None, ""


def find_first_startswith(sheet: Worksheet, prefixes: list[str], col: int = 2) -> tuple[Any | None, str]:
    for prefix in prefixes:
        value, original = find_startswith(sheet, prefix, col)
        if value is not None:
            return value, original
    return None, ""


def to_money(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value) * 1_000_000.0


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def add_metric(metrics: dict[str, XlsxMetric], key: str, value: float | None, method: str) -> None:
    metrics[key] = XlsxMetric(value, method)


def add_money_metric(metrics: dict[str, XlsxMetric], key: str, sheet_name: str, value: Any, label: str) -> None:
    add_metric(metrics, key, to_money(value), f"{sheet_name}: {label}")


def parse_outstanding_shares(sheet: Worksheet) -> XlsxMetric:
    for normalized, original, _ in worksheet_rows(sheet):
        if not normalized.startswith("common stock and paid-in capital"):
            continue
        match = re.search(r"outstanding\s+([0-9,]+)", original, re.IGNORECASE)
        if match:
            return XlsxMetric(float(match.group(1).replace(",", "")) * 1_000_000.0, f"Balance Sheets: {original}")
    return XlsxMetric(None, "Balance Sheets: missing common stock outstanding label")


def sum_available(values: list[float | None]) -> float | None:
    found = [value for value in values if value is not None]
    if not found:
        return None
    return sum(found)


def fiscal_report_period(fiscal_year: int, fiscal_quarter: int) -> str:
    if fiscal_quarter == 1:
        return date(fiscal_year - 1, 9, 30).isoformat()
    if fiscal_quarter == 2:
        return date(fiscal_year - 1, 12, 31).isoformat()
    if fiscal_quarter == 3:
        return date(fiscal_year, 3, 31).isoformat()
    if fiscal_quarter == 4:
        return date(fiscal_year, 6, 30).isoformat()
    raise ValueError(f"Unexpected Microsoft fiscal quarter: {fiscal_quarter}")


def calendar_quarter(report_period: str) -> str:
    report_date = datetime.strptime(report_period, "%Y-%m-%d")
    return f"{report_date.year}Q{((report_date.month - 1) // 3) + 1}"


def xlsx_file_info(path: Path) -> dict[str, Any] | None:
    match = re.search(r"FinancialStatementFY(?P<year>\d{2})Q(?P<quarter>[1-4])\.xlsx$", path.name, re.IGNORECASE)
    if not match:
        return None
    fiscal_year = 2000 + int(match.group("year"))
    fiscal_quarter = int(match.group("quarter"))
    report_period = fiscal_report_period(fiscal_year, fiscal_quarter)
    return {
        "fiscal_year": fiscal_year,
        "fiscal_quarter": fiscal_quarter,
        "fiscal_period": f"{fiscal_year}-Q{fiscal_quarter}",
        "report_period": report_period,
        "quarter": calendar_quarter(report_period),
        "path": str(path),
    }


def extract_workbook_metrics(path: Path) -> dict[str, XlsxMetric]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    income = workbook["Income Statements"]
    balance = workbook["Balance Sheets"]
    cash_flow = workbook["Cash Flows"]
    metrics: dict[str, XlsxMetric] = {}

    for key, label in {
        "cash_and_equivalents": "Cash and cash equivalents",
        "current_investments": "Short-term investments",
        "current_assets": "Total current assets",
        "inventory": "Inventories",
        "total_assets": "Total assets",
        "current_liabilities": "Total current liabilities",
        "current_debt": "Current portion of long-term debt",
        "trade_and_non_trade_payables": "Accounts payable",
        "deferred_revenue": "Short-term unearned revenue",
        "non_current_debt": "Long-term debt",
        "tax_liabilities": "Long-term income taxes",
        "total_liabilities": "Total liabilities",
        "shareholders_equity": "Total stockholders' equity",
        "retained_earnings": "Retained earnings",
        "accumulated_other_comprehensive_income": "Accumulated other comprehensive loss",
    }.items():
        value, source_label = find_exact(balance, label)
        add_money_metric(metrics, key, "Balance Sheets", value, source_label or label)

    for key, prefix in {
        "trade_and_non_trade_receivables": "Accounts receivable, net",
        "property_plant_and_equipment": "Property and equipment, net",
    }.items():
        value, source_label = find_startswith(balance, prefix)
        add_money_metric(metrics, key, "Balance Sheets", value, source_label or prefix)

    value, source_label = find_first_exact(balance, ["Equity and other investments", "Equity investments"])
    add_money_metric(metrics, "non_current_investments", "Balance Sheets", value, source_label or "Equity and other investments")

    goodwill = to_money(find_exact(balance, "Goodwill")[0]) or 0.0
    intangibles = to_money(find_exact(balance, "Intangible assets, net")[0]) or 0.0
    add_metric(metrics, "goodwill_and_intangible_assets", goodwill + intangibles, "Balance Sheets: Goodwill + Intangible assets, net")

    metrics["outstanding_shares"] = parse_outstanding_shares(balance)
    current_investments = metrics["current_investments"].value or 0.0
    non_current_investments = metrics["non_current_investments"].value or 0.0
    add_metric(metrics, "investments", current_investments + non_current_investments, "Balance Sheets: Short-term investments + equity/other investments")
    add_metric(
        metrics,
        "non_current_assets",
        (metrics["total_assets"].value or 0.0) - (metrics["current_assets"].value or 0.0),
        "Balance Sheets: Total assets - total current assets",
    )
    add_metric(
        metrics,
        "non_current_liabilities",
        (metrics["total_liabilities"].value or 0.0) - (metrics["current_liabilities"].value or 0.0),
        "Balance Sheets: Total liabilities - total current liabilities",
    )
    long_term_unearned = to_money(find_exact(balance, "Long-term unearned revenue")[0]) or 0.0
    add_metric(
        metrics,
        "deposit_liabilities",
        (metrics["deferred_revenue"].value or 0.0) + long_term_unearned,
        "Balance Sheets: Short-term unearned revenue + long-term unearned revenue",
    )
    add_metric(
        metrics,
        "total_debt",
        (metrics["current_debt"].value or 0.0) + (metrics["non_current_debt"].value or 0.0),
        "Balance Sheets: Current portion of long-term debt + long-term debt",
    )

    for key, label in {
        "revenue": "Total revenue",
        "cost_of_revenue": "Total cost of revenue",
        "gross_profit": "Gross margin",
        "research_and_development": "Research and development",
        "operating_income": "Operating income",
        "income_tax_expense": "Provision for income taxes",
        "net_income": "Net income",
    }.items():
        value, source_label = find_exact(income, label)
        add_money_metric(metrics, key, "Income Statements", value, source_label or label)

    sales_marketing = to_money(find_exact(income, "Sales and marketing")[0]) or 0.0
    general_admin = to_money(find_exact(income, "General and administrative")[0]) or 0.0
    r_and_d = metrics["research_and_development"].value or 0.0
    add_metric(metrics, "selling_general_and_administrative_expenses", sales_marketing + general_admin, "Income Statements: Sales and marketing + general and administrative")
    add_metric(metrics, "operating_expense", r_and_d + sales_marketing + general_admin, "Income Statements: R&D + sales and marketing + general and administrative")
    add_metric(metrics, "net_income_common_stock", metrics["net_income"].value, "Income Statements: Net income")
    add_metric(metrics, "preferred_dividends_impact", 0.0, "Income Statements: Microsoft reports no preferred dividends in the XLSX income statement")
    add_metric(metrics, "consolidated_income", metrics["net_income"].value, "Income Statements: Net income")

    basic_values: list[Any] = []
    diluted_values: list[Any] = []
    for normalized, _, row in worksheet_rows(income):
        if normalized == "basic":
            basic_values.append(row[1])
        elif normalized == "diluted":
            diluted_values.append(row[1])
    if len(basic_values) >= 2:
        add_metric(metrics, "earnings_per_share", to_float(basic_values[0]), "Income Statements: Basic EPS")
        add_metric(metrics, "weighted_average_shares", to_money(basic_values[1]), "Income Statements: Basic weighted average shares in millions")
    if len(diluted_values) >= 2:
        add_metric(metrics, "earnings_per_share_diluted", to_float(diluted_values[0]), "Income Statements: Diluted EPS")
        add_metric(metrics, "weighted_average_shares_diluted", to_money(diluted_values[1]), "Income Statements: Diluted weighted average shares in millions")

    for key, label in {
        "net_income": "Net income",
        "depreciation_and_amortization": "Depreciation, amortization, and other",
        "share_based_compensation": "Stock-based compensation expense",
        "net_cash_flow_from_operations": "Net cash from operations",
        "capital_expenditure": "Additions to property and equipment",
        "effect_of_exchange_rate_changes": "Effect of foreign exchange rates on cash and cash equivalents",
        "change_in_cash_and_equivalents": "Net change in cash and cash equivalents",
        "ending_cash_balance": "Cash and cash equivalents, end of period",
    }.items():
        value, source_label = find_exact(cash_flow, label)
        add_money_metric(metrics, key, "Cash Flows", value, source_label or label)

    value, source_label = find_first_exact(cash_flow, ["Net cash used in investing", "Net cash from (used in) investing"])
    add_money_metric(metrics, "net_cash_flow_from_investing", "Cash Flows", value, source_label or "Net cash from/used in investing")
    value, source_label = find_first_exact(cash_flow, ["Net cash used in financing", "Net cash from (used in) financing"])
    add_money_metric(metrics, "net_cash_flow_from_financing", "Cash Flows", value, source_label or "Net cash from/used in financing")
    value, source_label = find_first_startswith(cash_flow, ["Acquisition of companies, net of cash acquired"])
    add_money_metric(metrics, "business_acquisitions_and_disposals", "Cash Flows", value, source_label or "Acquisition of companies")

    purchases = to_money(find_exact(cash_flow, "Purchases of investments")[0]) or 0.0
    maturities = to_money(find_exact(cash_flow, "Maturities of investments")[0]) or 0.0
    sales = to_money(find_exact(cash_flow, "Sales of investments")[0]) or 0.0
    add_metric(metrics, "investment_acquisitions_and_disposals", purchases + maturities + sales, "Cash Flows: Purchases + maturities + sales of investments")

    debt_values = [
        to_money(find_exact(cash_flow, label)[0])
        for label in [
            "Proceeds from issuance of debt",
            "Proceeds from issuance of debt, maturities of 90 days or less, net",
            "Proceeds from issuance (repayments) of debt, maturities of 90 days or less, net",
            "Repayments of debt, maturities of 90 days or less",
            "Repayments of debt",
        ]
    ]
    add_metric(metrics, "issuance_or_repayment_of_debt_securities", sum_available(debt_values), "Cash Flows: Sum of debt issuance and repayment lines")

    common_stock_issued = to_money(find_exact(cash_flow, "Common stock issued")[0]) or 0.0
    common_stock_repurchased = to_money(find_exact(cash_flow, "Common stock repurchased")[0]) or 0.0
    add_metric(metrics, "issuance_or_purchase_of_equity_shares", common_stock_issued + common_stock_repurchased, "Cash Flows: Common stock issued + common stock repurchased")
    value, source_label = find_exact(cash_flow, "Common stock cash dividends paid")
    add_money_metric(metrics, "dividends_and_other_cash_distributions", "Cash Flows", value, source_label or "Common stock cash dividends paid")
    operating_cash_flow = metrics["net_cash_flow_from_operations"].value
    capex = metrics["capital_expenditure"].value
    fcf = None if operating_cash_flow is None or capex is None else operating_cash_flow - abs(capex)
    add_metric(metrics, "free_cash_flow", fcf, "Cash Flows: Net cash from operations - abs(additions to property and equipment)")

    return metrics


def load_xlsx_financials_by_fiscal_period(xlsx_dir: Path) -> dict[str, dict[str, Any]]:
    results: dict[str, dict[str, Any]] = {}
    for path in sorted(xlsx_dir.rglob("FinancialStatement*.xlsx")):
        info = xlsx_file_info(path)
        if info is None:
            continue
        info["metrics"] = extract_workbook_metrics(path)
        results[info["fiscal_period"]] = info
    return results


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows([{field: row.get(field, "") for field in fields} for row in rows])


def source_value(row: dict[str, Any], metric: str) -> float | None:
    value = row.get(metric)
    if value in (None, ""):
        return None
    return float(value)


def tolerance_for(unit_kind: str) -> float:
    if unit_kind == "per_share":
        return PER_SHARE_TOLERANCE
    if unit_kind == "shares":
        return SHARE_TOLERANCE
    return MONEY_TOLERANCE


def validation_status(metric_def: MetricDef, source: float | None, xlsx: float | None, tolerance: float) -> str:
    if not metric_def.comparable:
        return "not_comparable"
    if source is None:
        return "missing_source"
    if xlsx is None:
        return "missing_xlsx"
    if abs(source - xlsx) <= tolerance:
        return "match"
    if metric_def.metric == "capital_expenditure" and abs(abs(source) - abs(xlsx)) <= tolerance:
        return "sign_mismatch"
    return "mismatch"


def validate(company_rows: list[dict[str, Any]], xlsx_by_fiscal_period: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for company_row in sorted(company_rows, key=lambda row: row["report_period"]):
        fiscal_period = company_row["fiscal_period"]
        xlsx_info = xlsx_by_fiscal_period.get(fiscal_period)

        for metric_def in METRICS:
            xlsx_metric = None if xlsx_info is None else xlsx_info["metrics"].get(metric_def.metric)
            xlsx_value = None if xlsx_metric is None else xlsx_metric.value
            method = "missing FinancialStatement XLSX" if xlsx_metric is None else xlsx_metric.method
            current_value = source_value(company_row, metric_def.metric)
            tolerance = tolerance_for(metric_def.unit_kind)
            difference = None if current_value is None or xlsx_value is None else current_value - xlsx_value
            pct_difference = None
            if difference is not None and xlsx_value not in (None, 0):
                pct_difference = difference / abs(xlsx_value)

            rows.append(
                {
                    "ticker": TICKER,
                    "quarter": company_row["quarter"],
                    "report_period": company_row["report_period"],
                    "fiscal_period": fiscal_period,
                    "mapped_xlsx_quarter": "" if xlsx_info is None else xlsx_info["quarter"],
                    "mapped_xlsx_report_period": "" if xlsx_info is None else xlsx_info["report_period"],
                    "statement": metric_def.statement,
                    "metric": metric_def.metric,
                    "source_value": current_value,
                    "xlsx_value": xlsx_value,
                    "difference": difference,
                    "pct_difference": pct_difference,
                    "tolerance": tolerance,
                    "status": validation_status(metric_def, current_value, xlsx_value, tolerance),
                    "xlsx_method": method,
                    "xlsx_file": "" if xlsx_info is None else xlsx_info["path"],
                    "note": metric_def.note,
                }
            )

    return rows


def summarize(rows: list[dict[str, Any]]) -> dict[str, Any]:
    status_counts = Counter(row["status"] for row in rows)
    by_metric: dict[str, Counter[str]] = {}
    by_statement: dict[str, Counter[str]] = {}
    largest_mismatches = []

    for row in rows:
        by_metric.setdefault(row["metric"], Counter())
        by_metric[row["metric"]][row["status"]] += 1
        by_statement.setdefault(row["statement"], Counter())
        by_statement[row["statement"]][row["status"]] += 1
        if row["status"] in {"mismatch", "sign_mismatch"} and row["difference"] is not None:
            largest_mismatches.append(row)

    largest_mismatches = sorted(largest_mismatches, key=lambda row: abs(row["difference"]), reverse=True)[:30]
    core_metrics = [
        "cash_and_equivalents",
        "current_investments",
        "investments",
        "net_cash_flow_from_operations",
        "capital_expenditure",
        "free_cash_flow",
        "current_debt",
        "non_current_debt",
        "total_debt",
        "revenue",
        "operating_income",
        "net_income",
    ]

    return {
        "generated_at": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "ticker": TICKER,
        "rows": len(rows),
        "status_counts": dict(status_counts),
        "by_statement": {statement: dict(counts) for statement, counts in sorted(by_statement.items())},
        "by_metric": {metric: dict(counts) for metric, counts in sorted(by_metric.items())},
        "core_metric_status": {
            metric: dict(Counter(row["status"] for row in rows if row["metric"] == metric))
            for metric in core_metrics
        },
        "largest_mismatches": [
            {
                "quarter": row["quarter"],
                "fiscal_period": row["fiscal_period"],
                "metric": row["metric"],
                "status": row["status"],
                "source_value": row["source_value"],
                "xlsx_value": row["xlsx_value"],
                "difference": row["difference"],
                "xlsx_method": row["xlsx_method"],
                "note": row["note"],
            }
            for row in largest_mismatches
        ],
    }


def format_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return f"{value:,}"


def write_readme(output_dir: Path, summary: dict[str, Any]) -> None:
    status_lines = "\n".join(f"- {status}: {count}" for status, count in sorted(summary["status_counts"].items()))
    core_lines = "\n".join(f"- {metric}: {counts}" for metric, counts in summary["core_metric_status"].items())
    mismatch_lines = "\n".join(
        f"- {row['quarter']} {row['metric']} ({row['status']}): source={format_number(row['source_value'])}, xlsx={format_number(row['xlsx_value'])}, diff={format_number(row['difference'])}"
        for row in summary["largest_mismatches"][:12]
    )
    text = f"""# MSFT FinancialStatement XLSX Validation

Validation of stored MSFT financial data against locally archived Microsoft `FinancialStatement*.xlsx` files.

Generated: {summary['generated_at']}

## Period Mapping

Microsoft fiscal quarters are mapped to calendar quarter labels by report date. Example: `FY23Q3` is fiscal period `2023-Q3`, report period `2023-03-31`, and dataset quarter `2023Q1`.

## Status Counts

{status_lines}

## Core Metrics

{core_lines}

## Largest Mismatches

{mismatch_lines or "- None"}

## Files

- `validation_rows.csv` and `validation_rows.json`: metric-level comparisons.
- `summary.json`: aggregate counts and largest mismatches.
"""
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "README.md").write_text(text, encoding="utf-8")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--company-path", type=Path, default=DEFAULT_COMPANY_PATH)
    parser.add_argument("--xlsx-dir", type=Path, default=DEFAULT_XLSX_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    company_rows = read_json(args.company_path)
    xlsx_by_fiscal_period = load_xlsx_financials_by_fiscal_period(args.xlsx_dir)
    if not xlsx_by_fiscal_period:
        print(f"No Microsoft FinancialStatement XLSX files found in {args.xlsx_dir}", file=sys.stderr)
        return 1

    rows = validate(company_rows, xlsx_by_fiscal_period)
    summary = summarize(rows)
    fields = [
        "ticker",
        "quarter",
        "report_period",
        "fiscal_period",
        "mapped_xlsx_quarter",
        "mapped_xlsx_report_period",
        "statement",
        "metric",
        "source_value",
        "xlsx_value",
        "difference",
        "pct_difference",
        "tolerance",
        "status",
        "xlsx_method",
        "xlsx_file",
        "note",
    ]

    write_json(args.output_dir / "validation_rows.json", rows)
    write_csv(args.output_dir / "validation_rows.csv", rows, fields)
    write_json(args.output_dir / "summary.json", summary)
    write_readme(args.output_dir, summary)

    print(f"Validated {len(rows)} MSFT metric rows against {len(xlsx_by_fiscal_period)} FinancialStatement XLSX files")
    print(f"Status counts: {dict(sorted(summary['status_counts'].items()))}")
    print(f"Wrote validation output to {args.output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
