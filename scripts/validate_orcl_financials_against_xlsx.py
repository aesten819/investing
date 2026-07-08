#!/usr/bin/env python3
"""Validate stored ORCL financial rows against archived Oracle financial table XLSX files."""

from __future__ import annotations

import argparse
import calendar
import csv
import json
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

TICKER = "ORCL"
DEFAULT_COMPANY_PATH = Path("data/hyperscaler/companies/ORCL.json")
DEFAULT_XLSX_DIR = Path("data/hyperscaler/ir_documents/originals/ORCL/financial_tables_xlsx")
DEFAULT_OUTPUT_DIR = Path("data/hyperscaler/validation/orcl_financial_tables_xlsx")

MONEY_TOLERANCE = 1_000_000.0
SHARE_TOLERANCE = 1_000_000.0
PER_SHARE_TOLERANCE = 0.005


@dataclass(frozen=True)
class XlsxMetric:
    value: float | None
    method: str


@dataclass(frozen=True)
class MetricDef:
    metric: str
    statement: str
    unit_kind: str
    comparable: bool = True
    note: str = ""


BALANCE_METRICS = [
    MetricDef("total_assets", "balance_sheet", "money"),
    MetricDef("current_assets", "balance_sheet", "money"),
    MetricDef("cash_and_equivalents", "balance_sheet", "money"),
    MetricDef("inventory", "balance_sheet", "money", comparable=False, note="Oracle condensed financial tables do not present an inventory line."),
    MetricDef("current_investments", "balance_sheet", "money"),
    MetricDef("trade_and_non_trade_receivables", "balance_sheet", "money"),
    MetricDef("non_current_assets", "balance_sheet", "money"),
    MetricDef("property_plant_and_equipment", "balance_sheet", "money"),
    MetricDef("goodwill_and_intangible_assets", "balance_sheet", "money"),
    MetricDef("investments", "balance_sheet", "money", note="Mapped to marketable securities because Oracle's condensed table has no non-current investment line."),
    MetricDef("non_current_investments", "balance_sheet", "money", comparable=False, note="Oracle condensed financial tables do not present a non-current investment line."),
    MetricDef("outstanding_shares", "balance_sheet", "shares", comparable=False, note="Oracle financial tables provide weighted average shares, not balance-sheet shares outstanding."),
    MetricDef("tax_assets", "balance_sheet", "money"),
    MetricDef("total_liabilities", "balance_sheet", "money"),
    MetricDef("current_liabilities", "balance_sheet", "money"),
    MetricDef("current_debt", "balance_sheet", "money"),
    MetricDef("trade_and_non_trade_payables", "balance_sheet", "money"),
    MetricDef("deferred_revenue", "balance_sheet", "money"),
    MetricDef("deposit_liabilities", "balance_sheet", "money", comparable=False, note="Provider deposit_liabilities definition is not the same as Oracle current deferred revenues."),
    MetricDef("non_current_liabilities", "balance_sheet", "money"),
    MetricDef("non_current_debt", "balance_sheet", "money"),
    MetricDef("tax_liabilities", "balance_sheet", "money", comparable=False, note="Provider tax_liabilities maps to deferred tax liabilities in some periods; Oracle condensed table uses income taxes payable in others."),
    MetricDef("shareholders_equity", "balance_sheet", "money"),
    MetricDef("retained_earnings", "balance_sheet", "money", comparable=False, note="Oracle condensed financial tables do not present retained earnings."),
    MetricDef("accumulated_other_comprehensive_income", "balance_sheet", "money", comparable=False, note="Oracle condensed financial tables do not present accumulated OCI."),
    MetricDef("total_debt", "balance_sheet", "money"),
]

INCOME_METRICS = [
    MetricDef("revenue", "income_statement", "money"),
    MetricDef("cost_of_revenue", "income_statement", "money", note="Sum of Oracle cloud/software, hardware, and services cost rows."),
    MetricDef("gross_profit", "income_statement", "money", comparable=False, note="Oracle's condensed financial tables do not present gross profit directly."),
    MetricDef("operating_expense", "income_statement", "money", comparable=False, note="Provider operating_expense definition is not always the same as Oracle non-cost operating expense rows."),
    MetricDef("selling_general_and_administrative_expenses", "income_statement", "money", note="Sales and marketing plus general and administrative."),
    MetricDef("research_and_development", "income_statement", "money"),
    MetricDef("operating_income", "income_statement", "money"),
    MetricDef("interest_expense", "income_statement", "money", comparable=False, note="Provider interest_expense does not consistently match Oracle's interest expense row."),
    MetricDef("ebit", "income_statement", "money", comparable=False, note="Provider EBIT is not a directly reported Oracle financial table line."),
    MetricDef("income_tax_expense", "income_statement", "money"),
    MetricDef("net_income", "income_statement", "money"),
    MetricDef("net_income_common_stock", "income_statement", "money"),
    MetricDef("preferred_dividends_impact", "income_statement", "money"),
    MetricDef("consolidated_income", "income_statement", "money"),
    MetricDef("earnings_per_share", "income_statement", "per_share"),
    MetricDef("earnings_per_share_diluted", "income_statement", "per_share"),
    MetricDef("weighted_average_shares", "income_statement", "shares"),
    MetricDef("weighted_average_shares_diluted", "income_statement", "shares"),
]

CASH_FLOW_METRICS = [
    MetricDef("net_income", "cash_flow_statement", "money"),
    MetricDef("depreciation_and_amortization", "cash_flow_statement", "money", note="Depreciation plus amortization of intangible assets."),
    MetricDef("share_based_compensation", "cash_flow_statement", "money"),
    MetricDef("net_cash_flow_from_operations", "cash_flow_statement", "money"),
    MetricDef("capital_expenditure", "cash_flow_statement", "money", note="Oracle cash outflow sign from Capital expenditures."),
    MetricDef("business_acquisitions_and_disposals", "cash_flow_statement", "money"),
    MetricDef("investment_acquisitions_and_disposals", "cash_flow_statement", "money", note="Purchases plus proceeds from marketable securities and other investments."),
    MetricDef("net_cash_flow_from_investing", "cash_flow_statement", "money"),
    MetricDef("issuance_or_repayment_of_debt_securities", "cash_flow_statement", "money", note="Commercial paper plus senior notes, term loans, and other borrowing issuance/repayment lines."),
    MetricDef("issuance_or_purchase_of_equity_shares", "cash_flow_statement", "money", note="Common stock issuance, repurchases, tax withholding repurchases, and mandatory convertible preferred issuance when present."),
    MetricDef("dividends_and_other_cash_distributions", "cash_flow_statement", "money"),
    MetricDef("net_cash_flow_from_financing", "cash_flow_statement", "money"),
    MetricDef("change_in_cash_and_equivalents", "cash_flow_statement", "money"),
    MetricDef("effect_of_exchange_rate_changes", "cash_flow_statement", "money"),
    MetricDef("ending_cash_balance", "cash_flow_statement", "money"),
    MetricDef("free_cash_flow", "cash_flow_statement", "money", note="Calculated as net cash from operations minus absolute capital expenditures."),
]

METRICS = BALANCE_METRICS + INCOME_METRICS + CASH_FLOW_METRICS


def normalize_label(value: Any) -> str:
    text = str(value).replace("\xa0", " ").replace("\n", " ").replace("’", "'").replace("`", "'")
    return re.sub(r"\s+", " ", text).strip().lower()


def to_money(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value) * 1_000_000.0


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def add_metric(metrics: dict[str, XlsxMetric], key: str, value: float | None, method: str) -> None:
    metrics[key] = XlsxMetric(value, method)


def add_money_metric(metrics: dict[str, XlsxMetric], key: str, sheet_name: str, value: Any, label: str) -> None:
    add_metric(metrics, key, to_money(value), f"{sheet_name}: {label}")


def row_label(row: tuple[Any, ...], value_col: int, exact_label: str | None = None) -> str:
    cells = [cell for cell in row[: value_col - 1] if cell not in (None, "")]
    if exact_label:
        target = normalize_label(exact_label)
        for cell in cells:
            if normalize_label(cell) == target:
                return str(cell).replace("\n", " / ")
    return " | ".join(str(cell).replace("\n", " / ") for cell in cells)


def find_row(
    sheet: Worksheet,
    labels: list[str],
    value_col: int,
    *,
    startswith: bool = False,
) -> tuple[Any | None, str]:
    normalized_labels = [normalize_label(label) for label in labels]
    for row in sheet.iter_rows(values_only=True):
        for cell in row[: value_col - 1]:
            if cell in (None, ""):
                continue
            normalized = normalize_label(cell)
            matched = any(normalized.startswith(label) for label in normalized_labels) if startswith else normalized in normalized_labels
            if matched:
                original_label = row_label(row, value_col, str(cell))
                value = row[value_col - 1] if len(row) >= value_col else None
                return value, original_label
    return None, ""


def find_row_after(
    sheet: Worksheet,
    labels: list[str],
    value_col: int,
    marker: str,
    *,
    startswith: bool = False,
) -> tuple[Any | None, str]:
    normalized_marker = normalize_label(marker)
    normalized_labels = [normalize_label(label) for label in labels]
    marker_seen = False
    for row in sheet.iter_rows(values_only=True):
        row_cells = [cell for cell in row[: value_col - 1] if cell not in (None, "")]
        if not marker_seen:
            marker_seen = any(normalize_label(cell) == normalized_marker for cell in row_cells)
            continue
        for cell in row_cells:
            normalized = normalize_label(cell)
            matched = any(normalized.startswith(label) for label in normalized_labels) if startswith else normalized in normalized_labels
            if matched:
                original_label = row_label(row, value_col, str(cell))
                value = row[value_col - 1] if len(row) >= value_col else None
                return value, original_label
    return None, ""


def find_exact(sheet: Worksheet, label: str | list[str], value_col: int) -> tuple[Any | None, str]:
    labels = [label] if isinstance(label, str) else label
    return find_row(sheet, labels, value_col)


def find_startswith(sheet: Worksheet, prefix: str | list[str], value_col: int) -> tuple[Any | None, str]:
    prefixes = [prefix] if isinstance(prefix, str) else prefix
    return find_row(sheet, prefixes, value_col, startswith=True)


def money_from_exact(sheet: Worksheet, label: str | list[str], value_col: int) -> float | None:
    return to_money(find_exact(sheet, label, value_col)[0])


def money_from_startswith(sheet: Worksheet, prefix: str | list[str], value_col: int) -> float | None:
    return to_money(find_startswith(sheet, prefix, value_col)[0])


def money_from_exact_after(sheet: Worksheet, label: str | list[str], value_col: int, marker: str) -> float | None:
    labels = [label] if isinstance(label, str) else label
    return to_money(find_row_after(sheet, labels, value_col, marker)[0])


def sum_available(values: list[float | None]) -> float | None:
    found = [value for value in values if value is not None]
    if not found:
        return None
    return sum(found)


def zero_if_missing(value: float | None) -> float:
    return 0.0 if value is None else value


def extract_income_metrics(workbook: Any) -> dict[str, XlsxMetric]:
    sheet = workbook["QTD GAAP"]
    metrics: dict[str, XlsxMetric] = {}

    revenue, label = find_exact(sheet, "Total revenues", 4)
    add_money_metric(metrics, "revenue", "QTD GAAP", revenue, label or "Total revenues")

    cloud_cost = money_from_exact_after(sheet, ["Cloud and software", "Cloud services and license support"], 4, "Operating Expenses")
    hardware_cost = money_from_exact_after(sheet, "Hardware", 4, "Operating Expenses")
    services_cost = money_from_exact_after(sheet, "Services", 4, "Operating Expenses")
    cost_of_revenue = sum_available([cloud_cost, hardware_cost, services_cost])
    add_metric(metrics, "cost_of_revenue", cost_of_revenue, "QTD GAAP: Cloud/software cost + hardware cost + services cost")

    revenue_value = metrics["revenue"].value
    gross_profit = None if revenue_value is None or cost_of_revenue is None else revenue_value - cost_of_revenue
    add_metric(metrics, "gross_profit", gross_profit, "QTD GAAP: Total revenues - cost of revenue")

    sales_marketing = money_from_exact(sheet, "Sales and marketing", 4)
    general_admin = money_from_exact(sheet, "General and administrative", 4)
    research = money_from_exact(sheet, "Research and development", 4)
    amortization = money_from_exact(sheet, "Amortization of intangible assets", 4)
    acquisition_related = money_from_exact(sheet, "Acquisition related and other", 4)
    restructuring = money_from_exact(sheet, "Restructuring", 4)
    sga = sum_available([sales_marketing, general_admin])
    add_metric(metrics, "selling_general_and_administrative_expenses", sga, "QTD GAAP: Sales and marketing + general and administrative")
    add_metric(metrics, "research_and_development", research, "QTD GAAP: Research and development")
    add_metric(
        metrics,
        "operating_expense",
        sum_available([sales_marketing, general_admin, research, amortization, acquisition_related, restructuring]),
        "QTD GAAP: Non-cost operating expense rows",
    )

    for key, labels in {
        "operating_income": ["Operating income"],
        "interest_expense": ["Interest expense"],
        "net_income": ["Net income"],
    }.items():
        value, source_label = find_exact(sheet, labels, 4)
        add_money_metric(metrics, key, "QTD GAAP", value, source_label or labels[0])

    income_before_tax = money_from_exact(sheet, "Income before income taxes", 4)
    interest_expense = metrics["interest_expense"].value
    ebit = None if income_before_tax is None or interest_expense is None else income_before_tax + abs(interest_expense)
    add_metric(metrics, "ebit", ebit, "QTD GAAP: Income before income taxes + abs(interest expense)")

    tax_value, tax_label = find_startswith(
        sheet,
        ["Provision for income taxes", "(Provision for) benefit from income taxes", "Benefit from (provision for) income taxes"],
        4,
    )
    tax_money = to_money(tax_value)
    if tax_label and normalize_label(tax_label).startswith("benefit from"):
        tax_money = None if tax_money is None else -abs(tax_money)
    elif tax_label and normalize_label(tax_label).startswith("(provision for) benefit"):
        tax_money = None if tax_money is None else abs(tax_money)
    add_metric(metrics, "income_tax_expense", tax_money, f"QTD GAAP: {tax_label or 'income tax row'}")

    net_common, common_label = find_exact(sheet, "Net income available to common shareholders", 4)
    preferred_dividends, pref_label = find_exact(sheet, "Preferred stock dividends", 4)
    preferred_value = to_money(preferred_dividends)
    net_income_value = metrics["net_income"].value
    net_common_value = to_money(net_common)
    if net_common_value is None and net_income_value is not None and preferred_value is None:
        net_common_value = net_income_value
    add_metric(metrics, "net_income_common_stock", net_common_value, f"QTD GAAP: {common_label or 'Net income less preferred dividends when present'}")
    add_metric(metrics, "preferred_dividends_impact", preferred_value or 0.0, f"QTD GAAP: {pref_label or 'No preferred stock dividends line'}")
    add_metric(metrics, "consolidated_income", net_income_value, "QTD GAAP: Net income")

    basic_values: list[Any] = []
    diluted_values: list[Any] = []
    for row in sheet.iter_rows(values_only=True):
        if len(row) < 4:
            continue
        label = normalize_label(row[2]) if len(row) > 2 and row[2] not in (None, "") else ""
        if label == "basic":
            basic_values.append(row[3])
        elif label == "diluted":
            diluted_values.append(row[3])
    if len(basic_values) >= 2:
        add_metric(metrics, "earnings_per_share", to_float(basic_values[0]), "QTD GAAP: Basic EPS")
        add_metric(metrics, "weighted_average_shares", to_money(basic_values[1]), "QTD GAAP: Basic weighted average shares in millions")
    if len(diluted_values) >= 2:
        add_metric(metrics, "earnings_per_share_diluted", to_float(diluted_values[0]), "QTD GAAP: Diluted EPS")
        add_metric(metrics, "weighted_average_shares_diluted", to_money(diluted_values[1]), "QTD GAAP: Diluted weighted average shares in millions")

    return metrics


def extract_balance_metrics(workbook: Any) -> dict[str, XlsxMetric]:
    sheet = workbook["Balance Sheet"]
    metrics: dict[str, XlsxMetric] = {}

    for key, label in {
        "cash_and_equivalents": "Cash and cash equivalents",
        "current_investments": "Marketable securities",
        "trade_and_non_trade_receivables": "Trade receivables, net",
        "current_assets": "Total Current Assets",
        "property_plant_and_equipment": "Property, plant and equipment, net",
        "tax_assets": "Deferred tax assets",
        "non_current_assets": "Total Non-Current Assets",
        "total_assets": "Total assets",
        "current_debt": "Notes payable and other borrowings, current",
        "trade_and_non_trade_payables": "Accounts payable",
        "deferred_revenue": "Deferred revenues",
        "current_liabilities": "Total Current Liabilities",
        "non_current_debt": "Notes payable and other borrowings, non-current",
        "non_current_liabilities": "Total Non-Current Liabilities",
    }.items():
        value, source_label = find_exact(sheet, label, 5)
        add_money_metric(metrics, key, "Balance Sheet", value, source_label or label)

    shareholders_equity, equity_label = find_startswith(sheet, "Stockholders' Equity", 5)
    add_money_metric(metrics, "shareholders_equity", "Balance Sheet", shareholders_equity, equity_label or "Stockholders' Equity")

    goodwill = money_from_exact(sheet, ["Goodwill", "Goodwill, net"], 5)
    intangibles = money_from_exact(sheet, "Intangible assets, net", 5)
    add_metric(
        metrics,
        "goodwill_and_intangible_assets",
        zero_if_missing(goodwill) + zero_if_missing(intangibles),
        "Balance Sheet: Goodwill + intangible assets, net",
    )

    current_investments = metrics["current_investments"].value
    add_metric(metrics, "investments", current_investments, "Balance Sheet: Marketable securities")
    total_liabilities = sum_available([metrics["current_liabilities"].value, metrics["non_current_liabilities"].value])
    add_metric(metrics, "total_liabilities", total_liabilities, "Balance Sheet: Total current liabilities + total non-current liabilities")
    total_debt = sum_available([metrics["current_debt"].value, metrics["non_current_debt"].value])
    add_metric(metrics, "total_debt", total_debt, "Balance Sheet: Notes payable and other borrowings current + non-current")

    tax_liabilities = money_from_exact(sheet, ["Income taxes payable", "Deferred tax liabilities"], 5)
    add_metric(metrics, "tax_liabilities", tax_liabilities, "Balance Sheet: Income taxes payable or deferred tax liabilities")

    return metrics


def extract_cash_flow_ytd_metrics(workbook: Any) -> dict[str, XlsxMetric]:
    sheet = workbook["Cash Flow"]
    metrics: dict[str, XlsxMetric] = {}

    net_income = money_from_exact(sheet, "Net income", 4)
    depreciation = money_from_exact(sheet, "Depreciation", 4)
    amortization = money_from_exact(sheet, "Amortization of intangible assets", 4)
    share_comp = money_from_exact(sheet, "Stock-based compensation", 4)
    operating_cash = money_from_exact(sheet, "Net cash provided by operating activities", 4)
    capex = money_from_exact(sheet, "Capital expenditures", 4)
    acquisitions = money_from_exact(sheet, "Acquisitions, net of cash acquired", 4)
    purchase_investments = money_from_exact(sheet, "Purchases of marketable securities and other investments", 4)
    proceeds_investments = money_from_exact(sheet, "Proceeds from sales and maturities of marketable securities and other investments", 4)
    investing_cash = money_from_startswith(
        sheet,
        ["Net cash used for investing activities", "Net cash (used for) provided by investing activities"],
        4,
    )
    financing_cash = money_from_startswith(
        sheet,
        [
            "Net cash provided by financing activities",
            "Net cash provided by (used for) financing activities",
            "Net cash (used for) provided by financing activities",
            "Net cash used for financing activities",
        ],
        4,
    )
    dividends = money_from_exact(sheet, "Payments of dividends to stockholders", 4)
    effect_fx = money_from_exact(sheet, "Effect of exchange rate changes on cash and cash equivalents", 4)
    change_cash = money_from_startswith(
        sheet,
        [
            "Net increase in cash and cash equivalents",
            "Net increase (decrease) in cash and cash equivalents",
            "Net (decrease) increase in cash and cash equivalents",
            "Net decrease in cash and cash equivalents",
        ],
        4,
    )
    ending_cash = money_from_exact(sheet, "Cash and cash equivalents at end of period", 4)

    debt_values = [
        money_from_exact(sheet, label, 4)
        for label in [
            "Repayments of commercial paper",
            "Proceeds from issuances of commercial paper, net of repayments",
            "Proceeds from issuances of (repayments of) commercial paper, net",
            "Proceeds from (repayments of) commercial paper and other short-term financing, net",
            "Proceeds from issuances of senior notes and other borrowings, net of issuance costs",
            "Repayments of senior notes and other borrowings",
            "Proceeds from issuances of senior notes and term loan credit agreements, net of issuance costs",
            "Repayments of senior notes and term loan credit agreements",
            "Proceeds from issuances of senior notes, term loan credit agreements and other borrowings, net of issuance costs",
            "Repayments of senior notes, term loan credit agreements and other borrowings",
        ]
    ]
    equity_values = [
        money_from_exact(sheet, label, 4)
        for label in [
            "Proceeds from issuances of common stock",
            "Payments for repurchases of common stock",
            "Shares repurchased for tax withholdings upon vesting of restricted stock-based awards",
            "Proceeds from issuances of mandatory convertible preferred stock, net of issuance costs",
        ]
    ]

    add_metric(metrics, "net_income", net_income, "Cash Flow: Net income")
    add_metric(metrics, "depreciation_and_amortization", sum_available([depreciation, amortization]), "Cash Flow: Depreciation + amortization of intangible assets")
    add_metric(metrics, "share_based_compensation", share_comp, "Cash Flow: Stock-based compensation")
    add_metric(metrics, "net_cash_flow_from_operations", operating_cash, "Cash Flow: Net cash provided by operating activities")
    add_metric(metrics, "capital_expenditure", capex, "Cash Flow: Capital expenditures")
    add_metric(metrics, "business_acquisitions_and_disposals", acquisitions if acquisitions is not None else 0.0, "Cash Flow: Acquisitions, net of cash acquired")
    add_metric(metrics, "investment_acquisitions_and_disposals", sum_available([purchase_investments, proceeds_investments]), "Cash Flow: Purchases + proceeds of marketable securities and other investments")
    add_metric(metrics, "net_cash_flow_from_investing", investing_cash, "Cash Flow: Net cash used for investing activities")
    add_metric(metrics, "issuance_or_repayment_of_debt_securities", sum_available(debt_values), "Cash Flow: Debt issuance and repayment lines")
    add_metric(metrics, "issuance_or_purchase_of_equity_shares", sum_available(equity_values), "Cash Flow: Equity issuance, repurchase, tax withholding, and preferred issuance lines")
    add_metric(metrics, "dividends_and_other_cash_distributions", dividends, "Cash Flow: Payments of dividends to stockholders")
    add_metric(metrics, "net_cash_flow_from_financing", financing_cash, "Cash Flow: Net cash provided by/used for financing activities")
    add_metric(metrics, "change_in_cash_and_equivalents", change_cash, "Cash Flow: Net increase/decrease in cash and cash equivalents")
    add_metric(metrics, "effect_of_exchange_rate_changes", effect_fx, "Cash Flow: Effect of exchange rate changes on cash and cash equivalents")
    add_metric(metrics, "ending_cash_balance", ending_cash, "Cash Flow: Cash and cash equivalents at end of period")

    return metrics


def quarterize_cash_flow_metrics(results: dict[str, dict[str, Any]]) -> None:
    by_fiscal_year: dict[int, dict[int, dict[str, Any]]] = {}
    for info in results.values():
        by_fiscal_year.setdefault(info["fiscal_year"], {})[info["fiscal_quarter"]] = info

    point_in_time_keys = {"ending_cash_balance"}
    cash_flow_keys = [metric.metric for metric in CASH_FLOW_METRICS if metric.metric not in {"free_cash_flow", "net_income"}]
    for _, quarters in by_fiscal_year.items():
        previous_ytd: dict[str, XlsxMetric] | None = None
        for quarter in (1, 2, 3, 4):
            info = quarters.get(quarter)
            if info is None:
                previous_ytd = None
                continue
            ytd_metrics = info["cash_flow_ytd_metrics"]
            for key in cash_flow_keys:
                current = ytd_metrics.get(key)
                if current is None:
                    continue
                value = current.value
                method = current.method
                if key in point_in_time_keys:
                    info["metrics"][key] = current
                    continue
                if quarter > 1 and previous_ytd is not None:
                    previous = previous_ytd.get(key)
                    if value is not None and previous is not None and previous.value is not None:
                        value = value - previous.value
                        method = f"{current.method} less prior fiscal YTD"
                    else:
                        value = None
                elif quarter > 1:
                    value = None
                    method = f"{current.method}; missing prior fiscal YTD"
                info["metrics"][key] = XlsxMetric(value, method)
            operating_cash = info["metrics"].get("net_cash_flow_from_operations")
            capex = info["metrics"].get("capital_expenditure")
            fcf = None if operating_cash is None or capex is None or operating_cash.value is None or capex.value is None else operating_cash.value - abs(capex.value)
            info["metrics"]["free_cash_flow"] = XlsxMetric(fcf, "Cash Flow: Net cash provided by operating activities - abs(capital expenditures)")
            previous_ytd = ytd_metrics


def orcl_report_period(fiscal_year: int, fiscal_quarter: int) -> str:
    if fiscal_quarter == 1:
        return date(fiscal_year - 1, 8, 31).isoformat()
    if fiscal_quarter == 2:
        return date(fiscal_year - 1, 11, 30).isoformat()
    if fiscal_quarter == 3:
        return date(fiscal_year, 2, calendar.monthrange(fiscal_year, 2)[1]).isoformat()
    if fiscal_quarter == 4:
        return date(fiscal_year, 5, 31).isoformat()
    raise ValueError(f"Unexpected Oracle fiscal quarter: {fiscal_quarter}")


def xlsx_file_info(path: Path) -> dict[str, Any] | None:
    match = re.search(
        r"ORCL_(?P<quarter>\d{4}Q[1-4])_(?P<fiscal_year>\d{4})-Q(?P<fiscal_quarter>[1-4])_financial_tables_xlsx_.*\.xlsx$",
        path.name,
        re.IGNORECASE,
    )
    if not match:
        return None
    fiscal_year = int(match.group("fiscal_year"))
    fiscal_quarter = int(match.group("fiscal_quarter"))
    report_period = orcl_report_period(fiscal_year, fiscal_quarter)
    return {
        "fiscal_year": fiscal_year,
        "fiscal_quarter": fiscal_quarter,
        "fiscal_period": f"{fiscal_year}-Q{fiscal_quarter}",
        "report_period": report_period,
        "quarter": match.group("quarter"),
        "path": str(path),
    }


def extract_workbook_financials(path: Path) -> tuple[dict[str, XlsxMetric], dict[str, XlsxMetric]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    metrics: dict[str, XlsxMetric] = {}
    metrics.update(extract_balance_metrics(workbook))
    metrics.update(extract_income_metrics(workbook))
    cash_flow_ytd_metrics = extract_cash_flow_ytd_metrics(workbook)
    return metrics, cash_flow_ytd_metrics


def load_xlsx_financials_by_fiscal_period(xlsx_dir: Path) -> dict[str, dict[str, Any]]:
    results: dict[str, dict[str, Any]] = {}
    for path in sorted(xlsx_dir.glob("ORCL_*_financial_tables_xlsx_*.xlsx")):
        info = xlsx_file_info(path)
        if info is None:
            continue
        info["metrics"], info["cash_flow_ytd_metrics"] = extract_workbook_financials(path)
        results[info["fiscal_period"]] = info
    quarterize_cash_flow_metrics(results)
    return results


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows([{field: row.get(field, "") for field in fields} for row in rows])


def source_value(row: dict[str, Any], metric: str) -> float | None:
    value = row.get(metric)
    if value in (None, ""):
        return None
    return float(value)


def tolerance_for(unit_kind: str) -> float:
    if unit_kind == "per_share":
        return PER_SHARE_TOLERANCE
    if unit_kind == "shares":
        return SHARE_TOLERANCE
    return MONEY_TOLERANCE


def validation_status(metric_def: MetricDef, source: float | None, xlsx: float | None, tolerance: float) -> str:
    if not metric_def.comparable:
        return "not_comparable"
    if source is None:
        return "missing_source"
    if xlsx is None:
        return "missing_xlsx"
    if abs(source - xlsx) <= tolerance:
        return "match"
    if metric_def.metric == "capital_expenditure" and abs(abs(source) - abs(xlsx)) <= tolerance:
        return "sign_mismatch"
    return "mismatch"


def validate(company_rows: list[dict[str, Any]], xlsx_by_fiscal_period: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for company_row in sorted(company_rows, key=lambda row: row["report_period"]):
        fiscal_period = company_row["fiscal_period"]
        xlsx_info = xlsx_by_fiscal_period.get(fiscal_period)

        for metric_def in METRICS:
            xlsx_metric = None if xlsx_info is None else xlsx_info["metrics"].get(metric_def.metric)
            xlsx_value = None if xlsx_metric is None else xlsx_metric.value
            method = "missing Oracle financial tables XLSX" if xlsx_metric is None else xlsx_metric.method
            current_value = source_value(company_row, metric_def.metric)
            tolerance = tolerance_for(metric_def.unit_kind)
            difference = None if current_value is None or xlsx_value is None else current_value - xlsx_value
            pct_difference = None
            if difference is not None and xlsx_value not in (None, 0):
                pct_difference = difference / abs(xlsx_value)

            rows.append(
                {
                    "ticker": TICKER,
                    "quarter": company_row["quarter"],
                    "report_period": company_row["report_period"],
                    "fiscal_period": fiscal_period,
                    "mapped_xlsx_quarter": "" if xlsx_info is None else xlsx_info["quarter"],
                    "mapped_xlsx_report_period": "" if xlsx_info is None else xlsx_info["report_period"],
                    "statement": metric_def.statement,
                    "metric": metric_def.metric,
                    "source_value": current_value,
                    "xlsx_value": xlsx_value,
                    "difference": difference,
                    "pct_difference": pct_difference,
                    "tolerance": tolerance,
                    "status": validation_status(metric_def, current_value, xlsx_value, tolerance),
                    "xlsx_method": method,
                    "xlsx_file": "" if xlsx_info is None else xlsx_info["path"],
                    "note": metric_def.note,
                }
            )

    return rows


def summarize(rows: list[dict[str, Any]]) -> dict[str, Any]:
    status_counts = Counter(row["status"] for row in rows)
    by_metric: dict[str, Counter[str]] = {}
    by_statement: dict[str, Counter[str]] = {}
    largest_mismatches = []

    for row in rows:
        by_metric.setdefault(row["metric"], Counter())
        by_metric[row["metric"]][row["status"]] += 1
        by_statement.setdefault(row["statement"], Counter())
        by_statement[row["statement"]][row["status"]] += 1
        if row["status"] in {"mismatch", "sign_mismatch"} and row["difference"] is not None:
            largest_mismatches.append(row)

    largest_mismatches = sorted(largest_mismatches, key=lambda row: abs(row["difference"]), reverse=True)[:30]
    core_metrics = [
        "cash_and_equivalents",
        "current_investments",
        "net_cash_flow_from_operations",
        "capital_expenditure",
        "free_cash_flow",
        "current_debt",
        "non_current_debt",
        "total_debt",
        "revenue",
        "operating_income",
        "net_income",
    ]

    return {
        "generated_at": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "ticker": TICKER,
        "rows": len(rows),
        "status_counts": dict(status_counts),
        "by_statement": {statement: dict(counts) for statement, counts in sorted(by_statement.items())},
        "by_metric": {metric: dict(counts) for metric, counts in sorted(by_metric.items())},
        "core_metric_status": {
            metric: dict(Counter(row["status"] for row in rows if row["metric"] == metric))
            for metric in core_metrics
        },
        "largest_mismatches": [
            {
                "quarter": row["quarter"],
                "fiscal_period": row["fiscal_period"],
                "metric": row["metric"],
                "status": row["status"],
                "source_value": row["source_value"],
                "xlsx_value": row["xlsx_value"],
                "difference": row["difference"],
                "xlsx_method": row["xlsx_method"],
                "note": row["note"],
            }
            for row in largest_mismatches
        ],
    }


def format_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return f"{value:,}"


def write_readme(output_dir: Path, summary: dict[str, Any]) -> None:
    status_lines = "\n".join(f"- {status}: {count}" for status, count in sorted(summary["status_counts"].items()))
    core_lines = "\n".join(f"- {metric}: {counts}" for metric, counts in summary["core_metric_status"].items())
    mismatch_lines = "\n".join(
        f"- {row['quarter']} {row['metric']} ({row['status']}): source={format_number(row['source_value'])}, xlsx={format_number(row['xlsx_value'])}, diff={format_number(row['difference'])}"
        for row in summary["largest_mismatches"][:12]
    )
    text = f"""# ORCL Financial Tables XLSX Validation

Validation of stored ORCL financial data against locally archived Oracle official `financial_tables_xlsx` files.

Generated: {summary['generated_at']}

## Period Mapping

Oracle fiscal quarters are mapped to calendar quarter labels by report date. Example: fiscal `2026-Q3` maps to report period `2026-02-28` and dataset quarter `2026Q1`.

## Status Counts

{status_lines}

## Core Metrics

{core_lines}

## Largest Mismatches

{mismatch_lines or "- None"}

## Files

- `validation_rows.csv` and `validation_rows.json`: metric-level comparisons.
- `summary.json`: aggregate counts and largest mismatches.
"""
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "README.md").write_text(text, encoding="utf-8")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--company-path", type=Path, default=DEFAULT_COMPANY_PATH)
    parser.add_argument("--xlsx-dir", type=Path, default=DEFAULT_XLSX_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    company_rows = read_json(args.company_path)
    xlsx_by_fiscal_period = load_xlsx_financials_by_fiscal_period(args.xlsx_dir)
    if not xlsx_by_fiscal_period:
        print(f"No Oracle financial table XLSX files found in {args.xlsx_dir}", file=sys.stderr)
        return 1

    rows = validate(company_rows, xlsx_by_fiscal_period)
    summary = summarize(rows)
    fields = [
        "ticker",
        "quarter",
        "report_period",
        "fiscal_period",
        "mapped_xlsx_quarter",
        "mapped_xlsx_report_period",
        "statement",
        "metric",
        "source_value",
        "xlsx_value",
        "difference",
        "pct_difference",
        "tolerance",
        "status",
        "xlsx_method",
        "xlsx_file",
        "note",
    ]

    write_json(args.output_dir / "validation_rows.json", rows)
    write_csv(args.output_dir / "validation_rows.csv", rows, fields)
    write_json(args.output_dir / "summary.json", summary)
    write_readme(args.output_dir, summary)

    print(f"Validated {len(rows)} ORCL metric rows against {len(xlsx_by_fiscal_period)} Oracle financial table XLSX files")
    print(f"Status counts: {dict(sorted(summary['status_counts'].items()))}")
    print(f"Wrote validation output to {args.output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
